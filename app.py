import os
from datetime import datetime, date, timedelta, time
from flask import Flask, render_template, request, redirect, url_for, flash, session, send_file, jsonify
from functools import wraps
import sqlite3
import pandas as pd
from io import BytesIO  
import inspect
import aiohttp
import asyncio
from asgiref.wsgi import WsgiToAsgi
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.worksheet.dimensions import ColumnDimension

# Flask app setup
app = Flask(__name__)
app.secret_key = 'your-secret-key-here'  # Replace with a secure key

# Database path (relative to app.py in web_app directory)
DB_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'attendance.db')

# Telegram configuration (moved from telegram_utils.py)
TELEGRAM_BOT_TOKEN = '7311167814:AAFAHz3asyZSisZtgb9Tho79tWTkW4D4n24'  # Replace with your bot token
TELEGRAM_CHAT_ID = '-1002638304680'      # Replace with your chat ID

async def send_telegram_message(message):
    """
    Sends a message to a Telegram chat using the bot token and chat ID.
    
    Args:
        message (str): The message to send.
    
    Returns:
        bool: True if the message was sent successfully, False otherwise.
    """
    url = f"https://api.telegram.com/bot{TELEGRAM_BOT_TOKEN}/sendMessage"
    params = {
        'chat_id': TELEGRAM_CHAT_ID,
        'text': message
    }
    
    try:
        async with aiohttp.ClientSession() as session:
            async with session.get(url, params=params) as response:
                if response.status == 200:
                    print("Telegram message sent successfully.")
                    return True
                else:
                    print(f"Failed to send Telegram message: {response.status}")
                    return False
    except Exception as e:
        print(f"Error sending Telegram message: {e}")
        return False

async def send_telegram_report():
    """
    Generates an attendance report and sends it via Telegram.
    
    Returns:
        bool: True if the report was sent successfully, False otherwise.
    """
    today = date.today().isoformat()
    conn = get_db_connection()
    
    # Fetch today's attendance records
    records = conn.execute('''
        SELECT t.name, a.scan_time, a.scan_type
        FROM attendance a
        JOIN teachers t ON a.teacher_id = t.id
        WHERE a.date = ?
        ORDER BY a.scan_time
    ''', (today,)).fetchall()
    
    conn.close()
    
    # Generate the report message
    if not records:
        message = f"Attendance Report for {today}\nNo attendance records for today."
    else:
        message = f"Attendance Report for {today}\n\n"
        for record in records:
            # Convert scan_time to 12-hour format
            time_obj = datetime.strptime(record['scan_time'], '%H:%M:%S')
            formatted_time = time_obj.strftime('%I:%M %p')
            scan_type = record['scan_type'].replace('_', ' ').title()
            message += f"{record['name']} - {scan_type} at {formatted_time}\n"
    
    # Send the report via Telegram
    return await send_telegram_message(message)

def init_db():
    """
    Initializes the SQLite database by creating the required tables.
    Migrates the attendance table if it has an old schema.
    """
    # Ensure the directory exists
    db_dir = os.path.dirname(DB_PATH)
    if not os.path.exists(db_dir):
        os.makedirs(db_dir)
        print(f"Created database directory: {db_dir}")

    # Connect to the database
    conn = sqlite3.connect(DB_PATH)
    cursor = conn.cursor()

    # Create teachers table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS teachers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            class TEXT NOT NULL,
            card_id TEXT NOT NULL UNIQUE
        )
    ''')

    # Check if the attendance table exists and has the correct schema
    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='attendance'")
    table_exists = cursor.fetchone()

    if table_exists:
        # Check the schema of the existing attendance table
        cursor.execute("PRAGMA table_info(attendance)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'teacher_id' not in columns:
            print("Detected old schema for attendance table. Migrating...")
            # Rename the old table
            cursor.execute('ALTER TABLE attendance RENAME TO attendance_old')
            
            # Create the new attendance table with the correct schema
            cursor.execute('''
                CREATE TABLE attendance (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    teacher_id INTEGER NOT NULL,
                    date TEXT NOT NULL,
                    scan_time TEXT NOT NULL,
                    scan_type TEXT NOT NULL,
                    FOREIGN KEY (teacher_id) REFERENCES teachers(id)
                )
            ''')

            # Log a warning since we can't directly migrate data
            print("Warning: Existing attendance data cannot be migrated due to missing teacher_id column.")
            print("Please re-scan attendance or manually migrate data.")

            # Drop the old table
            cursor.execute('DROP TABLE attendance_old')
        else:
            print("Attendance table schema is up to date.")
    else:
        # Create the attendance table if it doesn't exist
        cursor.execute('''
            CREATE TABLE attendance (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                teacher_id INTEGER NOT NULL,
                date TEXT NOT NULL,
                scan_time TEXT NOT NULL,
                scan_type TEXT NOT NULL,
                FOREIGN KEY (teacher_id) REFERENCES teachers(id)
            )
        ''')

    # Check if scan_type column exists, and add it if not
    cursor.execute("PRAGMA table_info(attendance)")
    columns = [col[1] for col in cursor.fetchall()]
    if 'scan_type' not in columns:
        cursor.execute('ALTER TABLE attendance ADD COLUMN scan_type TEXT')
        cursor.execute('UPDATE attendance SET scan_type = "clock_in" WHERE scan_type IS NULL')

    # Debug: Print the final schema
    cursor.execute("PRAGMA table_info(attendance)")
    columns = cursor.fetchall()
    print("Final attendance table schema:")
    for col in columns:
        print(col)

    conn.commit()
    conn.close()
    print(f"Database initialized at: {DB_PATH}")

def get_db_connection():
    """
    Establishes a connection to the SQLite database with row factory set to return dict-like rows.
    
    Returns:
        sqlite3.Connection: Database connection object.
    """
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def login_required(f):
    """
    Decorator to ensure the user is logged in before accessing a route.
    Supports both sync and async functions.
    """
    @wraps(f)
    def wrap(*args, **kwargs):
        if 'username' not in session:
            flash('Please login to access this page.', 'error')
            return redirect(url_for('login'))
        if inspect.iscoroutinefunction(f):
            return f(*args, **kwargs)
        return f(*args, **kwargs)
    return wrap

# Define the global teacher list (70 entries after removing LOO POH HUA and MUHAMMAD IKMAL HAKIM B AHMAD SALMAN)
teacher_list = [
    {"Bil": "1", "Nama Guru": "AHMAD FARIZ BIN YUSOF", "Catatan": ""},
    {"Bil": "2", "Nama Guru": "NOORHAYATI BINTI MUSHROOM", "Catatan": ""},
    {"Bil": "3", "Nama Guru": "MOHAMAD PADZIL BIN SA’ALUDIN", "Catatan": ""},
    {"Bil": "4", "Nama Guru": "AZLINDA BINTI ALWI", "Catatan": ""},
    {"Bil": "5", "Nama Guru": "GHAZALI BIN RASID", "Catatan": ""},
    {"Bil": "6", "Nama Guru": "BADARIAH BINTI ABD GHANI", "Catatan": ""},
    {"Bil": "7", "Nama Guru": "BADROL HISHAM BIN ABDULLAH ISHAK", "Catatan": ""},
    {"Bil": "8", "Nama Guru": "NONI BINTI MD KARIM", "Catatan": ""},
    {"Bil": "9", "Nama Guru": "NORHAPIZA BINTI SULAIMAN", "Catatan": ""},
    {"Bil": "10", "Nama Guru": "HAZIELA BINTI MOHAMAD", "Catatan": ""},
    {"Bil": "11", "Nama Guru": "MEOR NOR NIKMAN BIN ABDUL RASHID", "Catatan": ""},
    {"Bil": "12", "Nama Guru": "ABDUL KADIR JALANI BIN AHMAD", "Catatan": ""},
    {"Bil": "13", "Nama Guru": "AHMAD BIN HJ SHAHADAN", "Catatan": ""},
    {"Bil": "14", "Nama Guru": "AIMAN FAYYADH BIN MOHAMAD ZIYAD", "Catatan": ""},
    {"Bil": "15", "Nama Guru": "AIMI DALILA BINTI AHMAD YATIM", "Catatan": ""},
    {"Bil": "16", "Nama Guru": "ANG SEONG CHUAN", "Catatan": ""},
    {"Bil": "17", "Nama Guru": "ASNIDAR BINTI MOHAMED ARIFF", "Catatan": ""},
    {"Bil": "18", "Nama Guru": "CHUAH SWEA YING", "Catatan": ""},
    {"Bil": "19", "Nama Guru": "DINESHWARY A/P PALANI", "Catatan": ""},
    {"Bil": "20", "Nama Guru": "FADZILAH BINTI IBRAHIM", "Catatan": ""},
    {"Bil": "21", "Nama Guru": "IBNOR REZA BIN YUSOF", "Catatan": ""},
    {"Bil": "22", "Nama Guru": "ILANGGO A/L KANAN", "Catatan": ""},
    {"Bil": "23", "Nama Guru": "KHAIRULNISYAH BINTI CHE RAHIM", "Catatan": ""},
    {"Bil": "24", "Nama Guru": "LIM SHOK ING", "Catatan": ""},
    {"Bil": "25", "Nama Guru": "MAGES A/P NAGAPPAN", "Catatan": ""},
    {"Bil": "26", "Nama Guru": "MAS AMIRUL SYAHMI BIN MASWIRA", "Catatan": ""},
    {"Bil": "27", "Nama Guru": "MOHD FAUZI BIN YUSOFF", "Catatan": ""},
    {"Bil": "28", "Nama Guru": "MOHD ZAMRI BIN ABDUL RAHMAN", "Catatan": ""},
    {"Bil": "29", "Nama Guru": "MUHAMAD AZWAR AFIFI BIN AZMAN", "Catatan": ""},
    {"Bil": "30", "Nama Guru": "MUHAMAD SUKRI BIN MOHAMED ZIN", "Catatan": ""},
    {"Bil": "31", "Nama Guru": "MUHAMMAD AIMAN BIN KHAIRUL ANUAR", "Catatan": ""},
    {"Bil": "32", "Nama Guru": "MUHAMMAD ZULFAZLI BIN ANUWAR", "Catatan": ""},
    {"Bil": "33", "Nama Guru": "MUHAMMAD ZULHILMI BIN BAAID", "Catatan": ""},
    {"Bil": "34", "Nama Guru": "MUNIROH BINTI MUSTAFFA", "Catatan": ""},
    {"Bil": "35", "Nama Guru": "NOOR AZLINA BINTI MD SAAD", "Catatan": ""},
    {"Bil": "36", "Nama Guru": "NOORA BINTI NOORDIN", "Catatan": ""},
    {"Bil": "37", "Nama Guru": "NOR AIYYUHAL BINTI JEMALI", "Catatan": ""},
    {"Bil": "38", "Nama Guru": "NOR ATIQAH BINTI MOHAMMAD ALI", "Catatan": ""},
    {"Bil": "39", "Nama Guru": "NOR HIDAYATUL AZALINA MOHD HARUN", "Catatan": ""},
    {"Bil": "40", "Nama Guru": "NOR MELATI BINTI YAKOP", "Catatan": ""},
    {"Bil": "41", "Nama Guru": "NOR SHAZLEEN BINTI MD AMIN", "Catatan": ""},
    {"Bil": "42", "Nama Guru": "NORAIN BINTI MOHD MARZUKI", "Catatan": ""},
    {"Bil": "43", "Nama Guru": "NORAZLIZA BINTI PUTEH", "Catatan": ""},
    {"Bil": "44", "Nama Guru": "NORLAILI BINTI ROSSELI", "Catatan": ""},
    {"Bil": "45", "Nama Guru": "NORULHUDA BINTI CHE EMBI", "Catatan": ""},
    {"Bil": "46", "Nama Guru": "NUR ATHIRAH BINTI MOHD YUSOF", "Catatan": ""},
    {"Bil": "47", "Nama Guru": "NUR IZATUL AQILAH BINTI IBRAHIN", "Catatan": ""},
    {"Bil": "48", "Nama Guru": "NUR SHAFIQAH BINTI MAHKHTAR", "Catatan": ""},
    {"Bil": "49", "Nama Guru": "NUR ZAIDA BINTI MOHD AZLAN", "Catatan": ""},
    {"Bil": "50", "Nama Guru": "NURNADHIRAH BINTI ABDUL RASHID", "Catatan": ""},
    {"Bil": "51", "Nama Guru": "NURSHAKINI BINTI MOHD BAHARDEN", "Catatan": ""},
    {"Bil": "52", "Nama Guru": "NURSHUHADA BINTI MOHAMAD JAZNI", "Catatan": ""},
    {"Bil": "53", "Nama Guru": "NURUL HUSNA BINTI HANAFI", "Catatan": ""},
    {"Bil": "54", "Nama Guru": "OLFFA BINTI SAMAT", "Catatan": ""},
    {"Bil": "55", "Nama Guru": "PRISNIE A/P PRAWING", "Catatan": ""},
    {"Bil": "56", "Nama Guru": "RAJA NOR ASHIKIN BINTI RAJA YEOP", "Catatan": "CUTI BERSALIN"},
    {"Bil": "57", "Nama Guru": "ROZY BIN AHMAD", "Catatan": ""},
    {"Bil": "58", "Nama Guru": "SANISAH BINTI MAT ALI", "Catatan": ""},
    {"Bil": "59", "Nama Guru": "SHAMINI A/P KAJUNATHAN", "Catatan": ""},
    {"Bil": "60", "Nama Guru": "SITI HAJAR BINTI NOPIAH", "Catatan": ""},
    {"Bil": "61", "Nama Guru": "SITI MAIZORA BINTI ABDUL RAHMAN", "Catatan": ""},
    {"Bil": "62", "Nama Guru": "SOFFIAN BIN HALIM", "Catatan": ""},
    {"Bil": "63", "Nama Guru": "SYAMSUL ANUAR BIN MOHD ISMAIL", "Catatan": ""},
    {"Bil": "64", "Nama Guru": "SYED MUHAMMAD AZRUL B SYED ABU HASSAN", "Catatan": ""},
    {"Bil": "65", "Nama Guru": "TAN MENG WEI", "Catatan": ""},
    {"Bil": "66", "Nama Guru": "WAN AHMAD UBAIDILLAH B WAN HANAPI", "Catatan": ""},
    {"Bil": "67", "Nama Guru": "YUSNIZAWATI BINTI YUSOF", "Catatan": ""},
    {"Bil": "68", "Nama Guru": "MUHAMMAD HANIF BIN HUSSIN", "Catatan": ""},
    {"Bil": "69", "Nama Guru": "NURFATNIN HADFINA BINTI MOHD HADAFI", "Catatan": ""},
    {"Bil": "70", "Nama Guru": "AZURA BINTI ISMAIL", "Catatan": ""},
    {"Bil": "71", "Nama Guru": "FAUZI BIN JAAFAR", "Catatan": ""},
    {"Bil": "72", "Nama Guru": "KHAIRUL ANUAR BIN SAID", "Catatan": ""},
    {"Bil": "73", "Nama Guru": "MAZWAN BIN BASHARUDIN", "Catatan": ""},
    {"Bil": "74", "Nama Guru": "NOORSAHIDA BINTI BAHARUDIN", "Catatan": ""},
    {"Bil": "75", "Nama Guru": "NORHIDAYAH BINTI MOHAMAD DZAKERI", "Catatan": ""},
    {"Bil": "76", "Nama Guru": "NURUL AIN HAFIZAH BINTI ABDUL MALEK", "Catatan": ""},
    {"Bil": "77", "Nama Guru": "ROSHITAH BINTI SARIFF", "Catatan": ""}
]

# Initialize the database when the app starts
init_db()

@app.route('/')
def index():
    if 'username' not in session:
        return redirect(url_for('login'))
    
    today = date.today().isoformat()
    conn = get_db_connection()
    
    # Total teachers
    total_teachers = conn.execute('SELECT COUNT(*) FROM teachers').fetchone()[0]
    
    # Today's attendance (distinct teachers who scanned today)
    today_attendance = conn.execute('''
        SELECT COUNT(DISTINCT teacher_id)
        FROM attendance
        WHERE date = ?
    ''', (today,)).fetchone()[0]
    
    # Pending scans (teachers who haven't scanned today)
    all_teachers = conn.execute('SELECT id FROM teachers').fetchall()
    scanned_teachers = conn.execute('''
        SELECT DISTINCT teacher_id
        FROM attendance
        WHERE date = ?
    ''', (today,)).fetchall()
    scanned_teacher_ids = {teacher['teacher_id'] for teacher in scanned_teachers}
    pending_scans = len([teacher for teacher in all_teachers if teacher['id'] not in scanned_teacher_ids])
    
    # Total records
    total_records = conn.execute('SELECT COUNT(*) FROM attendance').fetchone()[0]
    
    # Recent attendance (last 5 records)
    recent_attendance = conn.execute('''
        SELECT t.name AS teacher_name, a.scan_time AS timestamp, a.scan_type AS status
        FROM attendance a
        JOIN teachers t ON a.teacher_id = t.id
        ORDER BY a.date DESC, a.scan_time DESC
        LIMIT 5
    ''').fetchall()
    recent_attendance = [dict(record) for record in recent_attendance]
    
    # Convert timestamp to 12-hour format (hh:mm AM/PM)
    for record in recent_attendance:
        # Parse the scan_time (HH:MM:SS) and convert to 12-hour format
        time_obj = datetime.strptime(record['timestamp'], '%H:%M:%S')
        record['timestamp'] = time_obj.strftime('%I:%M %p')  # e.g., "10:49 AM"
        # Keep the raw scan_type ("clock_in" or "clock_out") as the status
    
    conn.close()
    
    return render_template('index.html',
                         total_teachers=total_teachers,
                         today_attendance=today_attendance,
                         pending_scans=pending_scans,
                         total_records=total_records,
                         recent_attendance=recent_attendance,
                         username=session['username'])

@app.route('/attendance')
@login_required
def attendance():
    conn = get_db_connection()
    today = date.today().isoformat()  # Get today's date in YYYY-MM-DD format

    # Fetch attendance records for today only, sorted by scan_time in descending order
    records = conn.execute('''
        SELECT a.id, t.name, t.class, t.id as teacher_id, a.date, a.scan_time, a.scan_type
        FROM attendance a
        JOIN teachers t ON a.teacher_id = t.id
        WHERE a.date = ?
        ORDER BY a.scan_time DESC
    ''', (today,)).fetchall()
    records = [dict(record) for record in records]

    # Log records to diagnose issues with scan_type
    for record in records:
        if not record['scan_type']:
            print(f"Warning: Record with ID {record['id']} for teacher {record['name']} on {record['date']} has no scan_type.")

    # Organize records by date (though it will only be today)
    daily_records = {}
    for record in records:
        record_date = record['date']
        if record_date not in daily_records:
            daily_records[record_date] = []
        daily_records[record_date].append(record)

    # For each date (only today), separate into clock_in and clock_out records
    clock_in_out_records = {}
    for record_date, date_records in daily_records.items():
        clock_in_out_records[record_date] = {'clock_in': [], 'clock_out': []}
        
        # Group records by teacher to determine their scan status
        teacher_scans = {}
        for record in date_records:
            teacher_id = record['teacher_id']
            if teacher_id not in teacher_scans:
                teacher_scans[teacher_id] = []
            teacher_scans[teacher_id].append(record)

        # Process each teacher's scans for the date
        for teacher_id, scans in teacher_scans.items():
            has_clock_out = any(scan['scan_type'] == 'clock_out' for scan in scans)
            has_clock_in = any(scan['scan_type'] == 'clock_in' for scan in scans)
            teacher_record = scans[0].copy()  # Base record for teacher info
            
            if has_clock_out:
                # Find the clock_out scan
                clock_out_scan = next((scan for scan in scans if scan['scan_type'] == 'clock_out'), None)
                if clock_out_scan:
                    # Convert scan_time to 12-hour format
                    time_obj = datetime.strptime(clock_out_scan['scan_time'], '%H:%M:%S')
                    teacher_record['scan_time'] = time_obj.strftime('%I:%M %p')  # e.g., "11:40 AM"
                    clock_in_out_records[record_date]['clock_out'].append(teacher_record)
                else:
                    print(f"Warning: Teacher {teacher_record['name']} on {record_date} has no clock_out scan but has_clock_out is True.")
            elif has_clock_in:
                # Find the clock_in scan
                clock_in_scan = next((scan for scan in scans if scan['scan_type'] == 'clock_in'), None)
                if clock_in_scan:
                    # Convert scan_time to 12-hour format
                    time_obj = datetime.strptime(clock_in_scan['scan_time'], '%H:%M:%S')
                    teacher_record['scan_time'] = time_obj.strftime('%I:%M %p')  # e.g., "08:47 AM"
                    clock_in_out_records[record_date]['clock_in'].append(teacher_record)
                else:
                    print(f"Warning: Teacher {teacher_record['name']} on {record_date} has no clock_in scan but has_clock_in is True.")
            else:
                # No valid clock_in or clock_out scans; log and skip
                print(f"Error: Teacher {teacher_record['name']} on {record_date} has no valid clock_in or clock_out scans. Scan types: {[scan['scan_type'] for scan in scans]}")

    # Fetch distinct classes for the dropdown
    teachers = conn.execute('SELECT DISTINCT class FROM teachers ORDER BY class').fetchall()
    teachers = [dict(teacher) for teacher in teachers]
    
    conn.close()
    
    return render_template('attendance.html', clock_in_out_records=clock_in_out_records, teachers=teachers)

@app.route('/export_excel', methods=["POST"])
@login_required
def export_excel():
    """
    Exports the attendance records to an Excel file with professional formatting.
    Includes columns: Bil, Nama Guru, Waktu Masuk, Waktu Keluar, Catatan.
    Matches teacher names exactly between database and document to assign Clock In/Out times.
    Filters by the selected class.
    """
    # Get selected date and class from the form submission
    selected_date = request.form.get('selected_date')
    selected_class = request.form.get('selected_class')

    if not selected_date or not selected_class:
        flash('Please select both a date and a class.', 'error')
        return redirect(url_for('attendance'))

    conn = get_db_connection()

    # Fetch teachers for the selected class from the database
    query = 'SELECT name FROM teachers WHERE class = ? ORDER BY name'
    teachers = conn.execute(query, (selected_class,)).fetchall()
    teachers = [dict(teacher) for teacher in teachers]

    # Fetch attendance records for the selected date and class
    attendance_query = '''
        SELECT t.name, t.class, a.date, a.scan_time, a.scan_type
        FROM attendance a
        JOIN teachers t ON a.teacher_id = t.id
        WHERE a.date = ? AND t.class = ?
        ORDER BY a.scan_time
    '''
    attendance_records = conn.execute(attendance_query, (selected_date, selected_class)).fetchall()
    attendance_records = [dict(record) for record in attendance_records]

    conn.close()

    # Group attendance records by teacher to get Clock In and Clock Out times
    teacher_attendance = {}
    for record in attendance_records:
        # Normalize teacher name to uppercase for case-insensitive matching
        teacher_name = record['name'].upper()
        if teacher_name not in teacher_attendance:
            teacher_attendance[teacher_name] = {'clock_in': None, 'clock_out': None}
        
        scan_time_str = f"{record['date']} {record['scan_time']}"
        scan_time = datetime.strptime(scan_time_str, "%Y-%m-%d %H:%M:%S")
        formatted_time = scan_time.strftime("%I:%M %p")

        if record['scan_type'] == 'clock_in':
            teacher_attendance[teacher_name]['clock_in'] = formatted_time
        elif record['scan_type'] == 'clock_out':
            teacher_attendance[teacher_name]['clock_out'] = formatted_time

    # Map Clock In/Out times by matching teacher names (case-insensitive)
    teacher_records = []
    matched_teachers = set()
    for teacher_entry in teacher_list:
        teacher_name = teacher_entry['Nama Guru']
        if not teacher_name:  # Skip empty entries
            continue
        
        # Normalize teacher name for matching
        teacher_name_upper = teacher_name.upper()
        # Only include teachers from the selected class
        if any(t['name'].upper() == teacher_name_upper for t in teachers):
            record = {
                'Bil': teacher_entry['Bil'],
                'Nama Guru': teacher_name,
                'Waktu Masuk': teacher_attendance.get(teacher_name_upper, {}).get('clock_in', ''),
                'Waktu Keluar': teacher_attendance.get(teacher_name_upper, {}).get('clock_out', ''),
                'Catatan': teacher_entry['Catatan']
            }
            teacher_records.append(record)
            matched_teachers.add(teacher_name_upper)

    # Log teachers in the database that didn't match any in the teacher_list
    db_teacher_names = {teacher['name'].upper() for teacher in teachers}
    unmatched_teachers = db_teacher_names - matched_teachers
    if unmatched_teachers:
        print("Unmatched teachers in database (check for name discrepancies):")
        for name in unmatched_teachers:
            print(f"- {name}")

    # Create DataFrame for Excel
    df = pd.DataFrame(teacher_records)

    # Create a BytesIO buffer to write the Excel file
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Attendance', startrow=2)  # Start at row 3 (1-based index) to leave space for title

        # Get the openpyxl workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Attendance']

        # Add a title above the table
        title = f"Rekod Kehadiran Guru {selected_date} - {selected_class}"
        worksheet['A1'] = title
        worksheet.merge_cells('A1:E1')  # Merge cells A1 to E1 for the title
        title_cell = worksheet['A1']
        title_cell.font = Font(name='Calibri', size=14, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Define styles
        header_font = Font(name='Calibri', size=11, bold=True)
        cell_font = Font(name='Calibri', size=11)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # Set column widths
        worksheet.column_dimensions['A'].width = 5   # Bil
        worksheet.column_dimensions['B'].width = 40  # Nama Guru
        worksheet.column_dimensions['C'].width = 15  # Waktu Masuk
        worksheet.column_dimensions['D'].width = 15  # Waktu Keluar
        worksheet.column_dimensions['E'].width = 20  # Catatan

        # Freeze the header row (row 3 in Excel, which is the header after the title)
        worksheet.freeze_panes = 'A4'

        # Apply formatting to the header row (row 3 in Excel)
        header_row = worksheet[3]  # 1-based index, row 3 is the header
        for cell in header_row:
            cell.font = header_font
            cell.alignment = center_alignment
            cell.border = border
            cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')

        # Apply formatting to data rows
        for row in worksheet.iter_rows(min_row=4, max_row=worksheet.max_row, min_col=1, max_col=5):
            for cell in row:
                cell.font = cell_font
                cell.border = border
                if cell.column_letter == 'B':
                    cell.alignment = left_alignment
                else:
                    cell.alignment = center_alignment

    # Seek to the beginning of the buffer
    output.seek(0)

    # Send the file to the user
    return send_file(
        output,
        as_attachment=True,
        download_name=f"rekod_kehadiran_{selected_date}_{selected_class}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route('/login', methods=['GET', 'POST'])
def login():
    """
    Handles user login with a simple username/password check.
    """
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        
        # Simple authentication (replace with proper user management in production)
        if username == 'tatbeng' and password == 'admin':
            session['username'] = username
            flash('Login successful!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    """
    Logs out the user by clearing the session.
    """
    session.pop('username', None)
    flash('You have been logged out.', 'success')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@login_required
def register():
    """
    Allows the admin to register a new teacher with name, class, and card ID.
    """
    if request.method == 'POST':
        name = request.form['name']
        class_name = request.form['class']
        card_id = request.form['card_id']
        
        conn = get_db_connection()
        conn.execute('INSERT INTO teachers (name, class, card_id) VALUES (?, ?, ?)',
                    (name, class_name, card_id))
        conn.commit()
        conn.close()
        
        flash('Teacher registered successfully!', 'success')
        return redirect(url_for('index'))
    
    return render_template('register.html')

# Helper function to check if the current time is between 6 AM and 9 AM
def is_clock_in_only_period():
    now = datetime.now().time()
    start_time = time(6, 0)  # 6 AM
    end_time = time(9, 0)    # 9 AM
    return start_time <= now <= end_time

@app.route('/scan', methods=['GET', 'POST'])
def scan():
    """
    Handles NFC card scans by recording the teacher's attendance.
    GET: Renders the scanning UI.
    POST: Processes the scanned card ID and records the attendance.
    Allows one "Clock In" and one "Clock Out" scan per day.
    Between 6 AM and 9 AM, only "Clock In" scans are allowed.
    """
    if request.method == 'POST':
        card_id = request.form['card_id'].strip()
        print(f"Received card_id: '{card_id}'")
        
        scan_time = datetime.now().strftime('%H:%M:%S')
        today = date.today().isoformat()
        
        conn = get_db_connection()
        teacher = conn.execute('SELECT id, name FROM teachers WHERE LOWER(card_id) = LOWER(?)', (card_id,)).fetchone()
        
        if not teacher:
            all_card_ids = conn.execute('SELECT card_id FROM teachers').fetchall()
            print(f"All card IDs in database: {[row['card_id'] for row in all_card_ids]}")
            conn.close()
            return {'status': 'error', 'message': 'Teacher not found'}, 404
        
        teacher_id = teacher['id']
        teacher_name = teacher['name']
        
        # Check existing scans for today
        existing_scans = conn.execute('''
            SELECT scan_type FROM attendance 
            WHERE teacher_id = ? AND date = ?
        ''', (teacher_id, today)).fetchall()
        
        # Determine scan type based on existing scans and current time
        if is_clock_in_only_period():
            # Between 6 AM and 9 AM, only allow "Clock In"
            if not existing_scans:
                # No scans today, record as "Clock In"
                scan_type = "clock_in"
            elif any(scan['scan_type'] == 'clock_in' for scan in existing_scans):
                # Already has a "Clock In", reject further scans
                conn.close()
                return {'status': 'warning', 'message': f'{teacher_name} has already clocked in today.'}, 200
            else:
                # This should not happen (no "Clock In" but trying to "Clock Out")
                conn.close()
                return {'status': 'warning', 'message': 'Clock out not allowed between 6 AM and 9 AM.'}, 200
        else:
            # After 9 AM, allow both "Clock In" and "Clock Out"
            if not existing_scans:
                # No scans today, this is a "Clock In"
                scan_type = "clock_in"
            elif len(existing_scans) == 1:
                # One scan exists, check if it's "Clock In"
                if any(scan['scan_type'] == 'clock_in' for scan in existing_scans):
                    # Has "Clock In", this must be "Clock Out"
                    scan_type = "clock_out"
                else:
                    # Has "Clock Out" (should not happen), reject
                    conn.close()
                    return {'status': 'warning', 'message': 'Invalid scan sequence. Please clock in first.'}, 200
            else:
                # Already have both "Clock In" and "Clock Out"
                conn.close()
                return {'status': 'warning', 'message': f'{teacher_name} has already completed both Clock In and Clock Out today.'}, 200
        
        # Insert the new scan
        conn.execute('INSERT INTO attendance (teacher_id, date, scan_time, scan_type) VALUES (?, ?, ?, ?)',
                    (teacher_id, today, scan_time, scan_type))
        conn.commit()
        conn.close()
        return {'status': 'success', 'message': f'{scan_type.replace("_", " ").title()} recorded for {teacher_name} at {scan_time}'}, 200
    
    return render_template('scan.html')

@app.route('/send-telegram-report', methods=['POST'])
@login_required
def send_telegram_report_route():
    """
    Route to send a Telegram report manually.
    """
    try:
        # Create a new event loop for this thread if needed
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        
        # Run the async coroutine using the new event loop
        success = loop.run_until_complete(send_telegram_report())
        
        # Close the event loop
        loop.close()
        
        if success:
            flash('Telegram report sent successfully!', 'success')
        else:
            flash('Failed to send Telegram report.', 'error')
    except Exception as e:
        flash(f'Error sending Telegram report: {str(e)}', 'error')
    return redirect(url_for('index'))

@app.route('/edit_db', methods=['GET', 'POST'])
@login_required
def edit_db():
    """
    Handles the edit database page for managing teachers.
    GET: Displays all teachers and forms for adding/updating/deleting.
    POST: Processes form submissions to add, update, delete, or bulk import teachers.
    """
    conn = get_db_connection()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        if action == 'add':
            name = request.form.get('name')
            class_name = request.form.get('class')
            card_id = request.form.get('card_id')
            if name and class_name and card_id:
                try:
                    conn.execute('INSERT INTO teachers (name, class, card_id) VALUES (?, ?, ?)',
                                (name, class_name, card_id))
                    conn.commit()
                    flash('Teacher added successfully!', 'success')
                except sqlite3.IntegrityError:
                    flash('Error: Card ID already exists.', 'error')
            else:
                flash('Please fill in all fields.', 'error')
        
        elif action == 'update':
            teacher_id = request.form.get('teacher_id')
            name = request.form.get('name')
            class_name = request.form.get('class')
            card_id = request.form.get('card_id')
            if teacher_id and (name or class_name or card_id):
                try:
                    updates = []
                    params = []
                    if name:
                        updates.append('name = ?')
                        params.append(name)
                    if class_name:
                        updates.append('class = ?')
                        params.append(class_name)
                    if card_id:
                        updates.append('card_id = ?')
                        params.append(card_id)
                    params.append(teacher_id)
                    if updates:
                        query = f'UPDATE teachers SET {", ".join(updates)} WHERE id = ?'
                        conn.execute(query, params)
                        conn.commit()
                        flash('Teacher updated successfully!', 'success')
                except sqlite3.IntegrityError:
                    flash('Error: Card ID already exists.', 'error')
            else:
                flash('Please provide at least one field to update.', 'error')
        
        elif action == 'delete':
            teacher_id = request.form.get('teacher_id')
            if teacher_id:
                # Delete associated attendance records first due to foreign key constraint
                conn.execute('DELETE FROM attendance WHERE teacher_id = ?', (teacher_id,))
                conn.execute('DELETE FROM teachers WHERE id = ?', (teacher_id,))
                conn.commit()
                flash('Teacher deleted successfully!', 'success')
            else:
                flash('Please select a teacher to delete.', 'error')
        
        elif action == 'bulk_import':
            # Insert teachers from teacher_list with placeholder card_ids
            for teacher in teacher_list:
                name = teacher['Nama Guru']
                if not name:  # Skip empty entries
                    continue
                # Check if teacher already exists to avoid duplicates
                existing = conn.execute('SELECT 1 FROM teachers WHERE name = ?', (name,)).fetchone()
                if not existing:
                    # Use a placeholder card_id (to be updated later)
                    placeholder_card_id = f"TBD_{teacher['Bil']}"
                    # Default class (can be updated later)
                    default_class = "Unknown"
                    try:
                        conn.execute('INSERT INTO teachers (name, class, card_id) VALUES (?, ?, ?)',
                                    (name, default_class, placeholder_card_id))
                    except sqlite3.IntegrityError:
                        print(f"Skipping teacher {name}: Card ID {placeholder_card_id} already exists.")
            conn.commit()
            flash('Teachers imported successfully! Please update their Card IDs.', 'success')
    
    # Fetch all teachers for display
    teachers = conn.execute('SELECT * FROM teachers ORDER BY name').fetchall()
    teachers = [dict(teacher) for teacher in teachers]
    
    conn.close()
    
    return render_template('edit_db.html', teachers=teachers)

@app.route('/statistics')
@login_required
def statistics():
    """
    Renders the statistics page showing attendance trends and analytics.
    """
    conn = get_db_connection()
    
    # Get attendance records for the last 7 days, counting distinct teachers per day
    today = date.today()
    seven_days_ago = (today - timedelta(days=7)).isoformat()
    
    daily_attendance = conn.execute('''
        SELECT date, COUNT(DISTINCT teacher_id) as count
        FROM attendance
        WHERE date >= ?
        GROUP BY date
        ORDER BY date DESC
    ''', (seven_days_ago,)).fetchall()
    
    # Get attendance by teacher (number of distinct days each teacher was present)
    # First, get all teachers
    all_teachers = conn.execute('SELECT id, name FROM teachers').fetchall()
    
    # Get attendance counts (distinct days) for teachers who have scans
    teacher_attendance_counts = conn.execute('''
        SELECT t.id, t.name, COUNT(DISTINCT a.date) as count
        FROM teachers t
        LEFT JOIN attendance a ON a.teacher_id = t.id
        WHERE a.date >= ? OR a.date IS NULL
        GROUP BY t.id, t.name
        ORDER BY count ASC  -- Order by ascending to show least attendance first
    ''', (seven_days_ago,)).fetchall()
    
    # Ensure all teachers are included, even those with 0 attendance
    teacher_attendance = []
    for teacher in all_teachers:
        teacher_id = teacher['id']
        attendance_record = next((record for record in teacher_attendance_counts if record['id'] == teacher_id), None)
        if attendance_record:
            teacher_attendance.append({'name': attendance_record['name'], 'count': attendance_record['count']})
        else:
            teacher_attendance.append({'name': teacher['name'], 'count': 0})
    
    # Get total records count (total number of scans)
    total_records = conn.execute('SELECT COUNT(*) as count FROM attendance').fetchone()['count']
    
    # Get today's attendance count (distinct teachers who scanned today)
    today_count = conn.execute('''
        SELECT COUNT(DISTINCT teacher_id) as count
        FROM attendance
        WHERE date = ?
    ''', (today.isoformat(),)).fetchone()['count']
    
    conn.close()
    
    # Prepare data for charts
    daily_labels = [record['date'] for record in daily_attendance]
    daily_counts = [record['count'] for record in daily_attendance]
    teacher_labels = [record['name'] for record in teacher_attendance]
    teacher_counts = [record['count'] for record in teacher_attendance]
    
    return render_template('statistics.html',
                         daily_labels=daily_labels,
                         daily_counts=daily_counts,
                         teacher_labels=teacher_labels,
                         teacher_counts=teacher_counts,
                         total_records=total_records,
                         today_count=today_count)

@app.route('/recent_attendance', methods=['GET'])
@login_required
def recent_attendance():
    """
    API endpoint to fetch the latest 5 attendance records in JSON format for real-time updates.
    Returns teacher name, timestamp, and status.
    """
    conn = get_db_connection()
    
    # Fetch the latest 5 attendance records
    recent_attendance = conn.execute('''
        SELECT t.name AS teacher_name, a.scan_time AS timestamp, a.scan_type AS status
        FROM attendance a
        JOIN teachers t ON a.teacher_id = t.id
        ORDER BY a.date DESC, a.scan_time DESC
        LIMIT 5
    ''').fetchall()
    recent_attendance = [dict(record) for record in recent_attendance]
    
    # Convert timestamp to 12-hour format (hh:mm AM/PM)
    for record in recent_attendance:
        time_obj = datetime.strptime(record['timestamp'], '%H:%M:%S')
        record['timestamp'] = time_obj.strftime('%I:%M %p')  # e.g., "11:14 AM"
        # Capitalize status for display (e.g., "clock_in" -> "CLOCK IN")
        record['status'] = record['status'].replace('_', ' ').upper()
    
    conn.close()
    
    return jsonify(recent_attendance)

# Wrap the Flask app in WsgiToAsgi for ASGI compatibility with uvicorn
asgi_app = WsgiToAsgi(app)

if __name__ == '__main__':
    app.run(debug=True)

if __name__ == '__main__':
    app.run(host='192.168.156.189', port=5000, debug=True)  # Accessible to network